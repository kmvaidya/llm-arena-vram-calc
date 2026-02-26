#!/usr/bin/env python3
"""
Arena.ai Open-Source LLM Leaderboard Enrichment with GPU VRAM Feasibility.

Scrapes the Arena.ai open-source text leaderboard, enriches each model with
parameter counts and VRAM estimates, and outputs an enriched CSV/XLSX showing
GPU deployment feasibility. Optionally updates the repo README.md with tables.

Usage:
    python enrich_arena.py                          # Scrape live from arena.ai
    python enrich_arena.py --input data.csv         # Use a pre-downloaded CSV
    python enrich_arena.py --update-readme          # Also update root README.md
"""

import argparse
import logging
import os
import re
import sys
from datetime import datetime, timezone

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Add parent directory to path so imports work when running directly
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from known_models import MODEL_OVERRIDES
from scrapers.arena_scraper import scrape_arena_leaderboard
from scrapers.aa_resolver import get_aa_models, resolve_from_aa, resolve_single_from_aa
from vram_calculator import (
    GPU_CONFIGS,
    GPU_DISPLAY_NAMES,
    GPUS_BY_VRAM,
    add_all_vram_and_gpu_columns,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger(__name__)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
REPO_ROOT = os.path.dirname(SCRIPT_DIR)

# Number of models to show directly in README (rest goes in <details>)
README_TOP_N = 50


# ---------------------------------------------------------------------------
# Parameter resolution
# ---------------------------------------------------------------------------

def resolve_from_overrides(model_name):
    """
    Strategy 1: Look up model in the small MODEL_OVERRIDES corrections dict.

    Only contains models where automated resolution gives wrong/no answers.
    Tries case-insensitive exact match first, then normalized substring match
    with safeguards to avoid false positives on distilled variants.
    """
    name_lower = model_name.lower()

    # Build case-insensitive lookup (lazy, rebuilt each call — dict is small)
    overrides_lower = {k.lower(): k for k in MODEL_OVERRIDES}

    # Exact match (case-insensitive)
    if name_lower in overrides_lower:
        orig_key = overrides_lower[name_lower]
        total, active, arch = MODEL_OVERRIDES[orig_key]
        return {
            "total_params_b": total,
            "active_params_b": active,
            "architecture": arch,
        }

    # Normalized substring match — normalize separators so "Command R+" matches
    # "command-r-plus" and "GLM-4.5-Air" matches "glm-4.5-air".
    def _normalize(s):
        return re.sub(r'[\s\-_]+', '-', s.lower()).replace('+', '-plus-').strip('-')

    name_norm = _normalize(model_name)

    # Skip substring matching if the name contains a size indicator (like
    # "Qwen3-8B" in "DeepSeek-R1-0528-Qwen3-8B") — that suggests a
    # different/distilled model. Let name parsing handle it instead.
    has_size_in_name = bool(re.search(
        r'(?:^|-)(\d+(?:\.\d+)?)b(?:-|$)',
        name_norm, re.IGNORECASE
    ))
    if has_size_in_name:
        return None

    best_match = None
    best_len = 0
    for key in MODEL_OVERRIDES:
        key_norm = _normalize(key)
        if key_norm in name_norm and len(key_norm) > best_len:
            best_match = key
            best_len = len(key_norm)

    if best_match:
        total, active, arch = MODEL_OVERRIDES[best_match]
        return {
            "total_params_b": total,
            "active_params_b": active,
            "architecture": arch,
        }

    return None


def resolve_from_name(model_name):
    """
    Strategy 2: Parse parameter count from model name using regex.

    Patterns handled:
        - Qwen3-30B-A3B       → total=30, active=3, moe
        - Llama-3.3-70B-Instruct → total=70, active=70, dense
        - Mistral-Small-2506-24B → total=24, active=24, dense
        - Gemma 3 27B IT       → total=27, active=27, dense
        - Mixtral-8x22B        → total=141 (8*22*0.8 rough), active=39, moe
        - Mixtral-8x7B         → total=47 (8*7*0.84), active=13, moe
    """
    # Special handling for Mixtral-style NxMB patterns
    mixtral_match = re.search(
        r'(?:^|[\s\-_])(\d+)x(\d+(?:\.\d+)?)\s*B(?:[\s\-_]|$)',
        model_name,
        re.IGNORECASE,
    )
    if mixtral_match:
        num_experts = int(mixtral_match.group(1))
        expert_size = float(mixtral_match.group(2))
        # MoE total ≈ num_experts * expert_size * overhead_factor (shared layers add ~20%)
        # Active ≈ 2 experts active (typical for Mixtral) + shared params
        total_b = round(num_experts * expert_size * 0.8, 1)  # rough estimate
        active_b = round(expert_size * 2 * 0.9, 1)  # 2 active experts typical
        return {
            "total_params_b": total_b,
            "active_params_b": active_b,
            "architecture": "moe",
        }

    # Look for active params first: A{N}B pattern (MoE indicator)
    active_match = re.search(
        r'(?:^|[\s\-_])A(\d+(?:\.\d+)?)\s*B(?:[\s\-_]|$)',
        model_name,
        re.IGNORECASE,
    )

    # Look for total size: {N}B pattern
    # Must be preceded by a separator and followed by B (not part of longer word)
    total_match = re.search(
        r'(?:^|[\s\-_])(\d+(?:\.\d+)?)\s*B(?:[\s\-_]|$)',
        model_name,
        re.IGNORECASE,
    )

    if not total_match:
        return None

    total_b = float(total_match.group(1))

    if active_match:
        active_b = float(active_match.group(1))
        return {
            "total_params_b": total_b,
            "active_params_b": active_b,
            "architecture": "moe",
        }

    return {
        "total_params_b": total_b,
        "active_params_b": total_b,
        "architecture": "dense",
    }


def resolve_model_params(model_name, use_network=True, aa_lookup=None):
    """
    Resolve parameter counts for a model using the priority chain.

    Priority:
        1. MODEL_OVERRIDES (corrections for AA mistakes, misleading names, etc.)
        2. Artificial Analysis cache (bulk data, covers most open-source models)
        3. Parse from model name (fast, no network — last resort)
        4. AA single-model scrape (if use_network, for models missing from bulk)
        5. UNKNOWN

    Args:
        model_name: Arena model name
        use_network: Whether to allow network requests (AA single-model fallback)
        aa_lookup: Pre-loaded AA lookup dict (pass for efficiency in batch calls)

    Returns:
        tuple of (params_dict, source_name)
    """
    # Strategy 1: Override corrections (highest priority — fixes AA mistakes)
    result = resolve_from_overrides(model_name)
    if result:
        return result, "override"

    # Strategy 2: Artificial Analysis cache (primary automated source)
    if aa_lookup is not None:
        try:
            result = resolve_from_aa(model_name, aa_lookup=aa_lookup)
            if result:
                return result, "artificial_analysis"
        except Exception as e:
            logger.debug(f"AA cache resolution failed for {model_name}: {e}")

    # Strategy 3: Parse from model name
    result = resolve_from_name(model_name)
    if result:
        return result, "name_parsing"

    # Strategy 4: AA single-model page scrape (expensive, last resort)
    if use_network:
        try:
            result = resolve_single_from_aa(model_name)
            if result:
                return result, "aa_single_scrape"
        except Exception as e:
            logger.debug(f"AA single scrape failed for {model_name}: {e}")

    return None, "UNKNOWN"


# ---------------------------------------------------------------------------
# XLSX formatting
# ---------------------------------------------------------------------------

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BOLD_FONT = Font(bold=True)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def write_xlsx(df, resolution_log, output_path):
    """Write enriched data to a formatted XLSX file with 3 sheets."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_full_table(df, writer)
        _write_gpu_summary(df, writer)
        _write_resolution_log(resolution_log, writer)
    logger.info(f"XLSX saved to: {output_path}")


def _write_full_table(df, writer):
    """Write the full enriched table with conditional formatting."""
    df.to_excel(writer, sheet_name="Full Table", index=False)
    ws = writer.sheets["Full Table"]

    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for row_idx in range(2, min(12, len(df) + 2)):
        for col_idx in range(1, len(df.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).font = BOLD_FONT

    fit_cols = [c for c in df.columns if c.startswith("fits_")]
    for col_name in fit_cols:
        col_idx = list(df.columns).index(col_name) + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is True:
                cell.fill = GREEN_FILL
                cell.value = "Yes"
            elif cell.value is False:
                cell.fill = RED_FILL
                cell.value = "No"

    for col_idx in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(20, len(df) + 2))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _write_gpu_summary(df, writer):
    """Write GPU summary sheet."""
    rows = _build_gpu_summary_rows(df)
    summary_df = pd.DataFrame(rows)
    summary_df.to_excel(writer, sheet_name="GPU Summary", index=False)

    ws = writer.sheets["GPU Summary"]
    for col_idx in range(1, len(summary_df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for col_idx in range(1, len(summary_df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(30, len(summary_df) + 2))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _write_resolution_log(resolution_log, writer):
    """Write resolution log sheet."""
    log_df = pd.DataFrame(resolution_log)
    log_df.to_excel(writer, sheet_name="Resolution Log", index=False)

    ws = writer.sheets["Resolution Log"]
    for col_idx in range(1, len(log_df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for col_idx in range(1, len(log_df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, min(30, len(log_df) + 2))
        )
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def _build_gpu_summary_rows(df):
    """Build GPU summary data for both XLSX and README."""
    rows = []
    for gpu_key, gpu_vram in GPUS_BY_VRAM:
        display = GPU_DISPLAY_NAMES[gpu_key]
        for precision in ["BF16", "FP8", "INT4"]:
            p_lower = precision.lower()
            serving_col = f"vram_{p_lower}_serving_gb"
            if serving_col not in df.columns:
                continue

            fits = df[df[serving_col].notna() & (df[serving_col] <= gpu_vram)]
            if not fits.empty:
                best = fits.iloc[0]
                note = ""
                if gpu_key == "RTX_PRO_6000" and precision == "FP8":
                    note = "FP8 requires software emulation"
                rows.append({
                    "GPU": f"{display} ({gpu_vram} GB)",
                    "Precision": precision,
                    "Best Rank": int(best["rank"]) if pd.notna(best["rank"]) else "?",
                    "Model": best["model_name"],
                    "Total Params (B)": best.get("total_params_b", "?"),
                    "Active Params (B)": best.get("active_params_b", "?"),
                    "Architecture": best.get("architecture", "?"),
                    "Serving VRAM (GB)": round(best[serving_col], 1),
                    "Note": note,
                })
            else:
                rows.append({
                    "GPU": f"{display} ({gpu_vram} GB)",
                    "Precision": precision,
                    "Best Rank": "N/A",
                    "Model": "No model fits",
                    "Total Params (B)": "",
                    "Active Params (B)": "",
                    "Architecture": "",
                    "Serving VRAM (GB)": "",
                    "Note": "",
                })
    return rows


# ---------------------------------------------------------------------------
# README generation
# ---------------------------------------------------------------------------

def generate_readme(df, resolution_counts, aa_is_stale=False):
    """
    Generate the full README.md content with:
    1. Best Model Per GPU — one table per precision
    2. Enriched Leaderboard — compact table with collapsible rest
    3. How It Works / Usage
    """
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    total = len(df)
    resolved = total - resolution_counts.get("UNKNOWN", 0)
    pct = (resolved / total * 100) if total > 0 else 0

    lines = []

    # Header
    lines.append("# LLM Arena VRAM Calculator")
    lines.append("")
    lines.append("Enriches the [Arena.ai](https://arena.ai/leaderboard/text?license=open-source) "
                 "open-source LLM leaderboard with parameter counts and VRAM estimates "
                 "for single-GPU deployment feasibility.")
    lines.append("")
    lines.append(f"> **Last updated:** {now} | "
                 f"**Models:** {total} | "
                 f"**Resolved:** {resolved} ({pct:.1f}%)")
    lines.append("")
    if aa_is_stale:
        lines.append("> **Warning:** AA data may be stale (RSC fetch failed, using cached data).")
        lines.append("")

    # Table 1: Best Model Per GPU — one sub-table per precision
    lines.append("## Best Model Per GPU")
    lines.append("")
    lines.append("Highest-ranked Arena model that fits on each single GPU "
                 "(includes 25% serving overhead for KV cache, activations, and framework).")
    lines.append("")
    for precision in ["BF16", "FP8", "INT4"]:
        lines.extend(_generate_gpu_precision_table(df, precision))
        lines.append("")

    # Table 2: Enriched Leaderboard
    lines.append("## Full Leaderboard")
    lines.append("")
    lines.extend(_generate_leaderboard_table(df))
    lines.append("")

    # How It Works section
    lines.append("## How It Works")
    lines.append("")
    lines.append("### VRAM Estimation")
    lines.append("")
    lines.append("| Precision | Bytes/Param | Example: 70B model |")
    lines.append("|-----------|-------------|---------------------|")
    lines.append("| BF16 | 2.0 | 140 GB weights, 175 GB serving |")
    lines.append("| FP8 | 1.0 | 70 GB weights, 87.5 GB serving |")
    lines.append("| INT4 | 0.5 | 35 GB weights, 43.8 GB serving |")
    lines.append("")
    lines.append("- **Serving VRAM** = weight VRAM x 1.25 "
                 "(25% overhead for KV cache, activations, framework)")
    lines.append("- For **MoE models**, VRAM uses **total** parameters "
                 "(all experts must be loaded)")
    lines.append("")
    lines.append("### Parameter Resolution")
    lines.append("")
    lines.append("Parameters are resolved via a priority chain:")
    lines.append("")
    lines.append("1. **Override corrections** — models with misleading names "
                 "or wrong AA data")
    lines.append("2. **[Artificial Analysis](https://artificialanalysis.ai)** "
                 "— cached bulk model database (primary source)")
    lines.append("3. **Name parsing** — regex extraction of `{N}B` and `A{N}B` "
                 "patterns")
    lines.append("")
    lines.append("### GPUs")
    lines.append("")
    lines.append("| GPU | VRAM | Architecture | Native FP8 |")
    lines.append("|-----|------|-------------|------------|")
    lines.append("| H100 SXM | 80 GB | Hopper | Yes |")
    lines.append("| RTX PRO 6000 | 96 GB | Ada Lovelace | No (software emulation) |")
    lines.append("| H200 SXM | 141 GB | Hopper | Yes |")
    lines.append("| B200 SXM | 180 GB | Blackwell | Yes |")
    lines.append("| B300 SXM | 288 GB | Blackwell Ultra | Yes |")
    lines.append("")

    # Usage section
    lines.append("## Usage")
    lines.append("")
    lines.append("```bash")
    lines.append("cd arena_enrichment")
    lines.append("pip install -r requirements.txt")
    lines.append("")
    lines.append("# Full pipeline (scrapes arena.ai live)")
    lines.append("python enrich_arena.py --update-readme")
    lines.append("")
    lines.append("# Use a pre-downloaded CSV")
    lines.append("python enrich_arena.py --input data.csv --update-readme")
    lines.append("")
    lines.append("# Skip network resolution (overrides + name parsing only)")
    lines.append("python enrich_arena.py --input data.csv --no-network --update-readme")
    lines.append("```")
    lines.append("")

    return "\n".join(lines)


def _generate_gpu_precision_table(df, precision):
    """Generate a compact GPU summary table for a single precision."""
    lines = []
    p_lower = precision.lower()
    serving_col = f"vram_{p_lower}_serving_gb"

    label = {"BF16": "BF16 (Full Precision)", "FP8": "FP8 (8-bit)",
             "INT4": "INT4 (4-bit)"}[precision]
    lines.append(f"### {label}")
    lines.append("")
    lines.append("| GPU | VRAM | Best Model | Arena Rank | Params | Arch | Serving VRAM |")
    lines.append("|-----|------|------------|------------|--------|------|--------------|")

    if serving_col not in df.columns:
        return lines

    for gpu_key, gpu_vram in GPUS_BY_VRAM:
        display = GPU_DISPLAY_NAMES[gpu_key]
        fits = df[df[serving_col].notna() & (df[serving_col] <= gpu_vram)]
        if not fits.empty:
            best = fits.iloc[0]
            rank = int(best["rank"]) if pd.notna(best["rank"]) else "?"
            name = best["model_name"]
            total = best.get("total_params_b")
            total_s = f"{total:g}B" if pd.notna(total) else "?"
            arch = (best.get("architecture") or "?")
            if pd.isna(arch):
                arch = "?"
            elif arch == "dense":
                arch = "Dense"
            elif arch == "moe":
                arch = "MoE"
            serving = f"{round(best[serving_col], 1)} GB"
            lines.append(f"| {display} | {gpu_vram} GB | {name} | #{rank} | {total_s} | {arch} | {serving} |")
        else:
            lines.append(f"| {display} | {gpu_vram} GB | *No model fits* | — | — | — | — |")

    return lines


def _generate_leaderboard_table(df):
    """Generate the enriched leaderboard markdown table."""
    lines = []

    header = "| Rank | Model | Score | Params (B) | Arch | VRAM BF16 | VRAM FP8 | VRAM INT4 | Fits on |"
    sep = "|------|-------|-------|------------|------|-----------|----------|-----------|---------|"

    def _format_row(row):
        rank = int(row["rank"]) if pd.notna(row.get("rank")) else "?"
        name = row.get("model_name", "?")
        score = int(row["arena_score"]) if pd.notna(row.get("arena_score")) else "?"
        total = row.get("total_params_b")
        active = row.get("active_params_b")
        # Show "total (active)" for MoE, just "total" for dense
        if pd.notna(total) and pd.notna(active) and total != active:
            params_s = f"{total:g} ({active:g})"
        elif pd.notna(total):
            params_s = f"{total:g}"
        else:
            params_s = "?"
        arch = (row.get("architecture") or "?").upper() if pd.notna(row.get("architecture")) else "?"
        if arch == "DENSE":
            arch = "Dense"
        elif arch == "MOE":
            arch = "MoE"
        bf16 = row.get("vram_bf16_serving_gb")
        bf16_s = f"{bf16:g}" if pd.notna(bf16) else "?"
        fp8 = row.get("vram_fp8_serving_gb")
        fp8_s = f"{fp8:g}" if pd.notna(fp8) else "?"
        int4 = row.get("vram_int4_serving_gb")
        int4_s = f"{int4:g}" if pd.notna(int4) else "?"
        # "Fits on" — show smallest single GPU at FP8, the most practical precision
        best_fp8 = row.get("best_gpu_fp8", "?")
        if best_fp8 == "MULTI-GPU":
            fits_s = "Multi-GPU"
        elif best_fp8 == "UNKNOWN":
            fits_s = "?"
        else:
            fits_s = f"{best_fp8} (FP8)"
        return f"| {rank} | {name} | {score} | {params_s} | {arch} | {bf16_s} | {fp8_s} | {int4_s} | {fits_s} |"

    # Top N rows shown directly
    top_df = df.head(README_TOP_N)
    rest_df = df.iloc[README_TOP_N:]

    lines.append(header)
    lines.append(sep)
    for _, row in top_df.iterrows():
        lines.append(_format_row(row))

    # Remaining rows in collapsible section
    if len(rest_df) > 0:
        lines.append("")
        lines.append(f"<details><summary>Show remaining {len(rest_df)} models</summary>")
        lines.append("")
        lines.append(header)
        lines.append(sep)
        for _, row in rest_df.iterrows():
            lines.append(_format_row(row))
        lines.append("")
        lines.append("</details>")

    return lines


# ---------------------------------------------------------------------------
# Console summary
# ---------------------------------------------------------------------------

def print_console_summary(df, resolution_counts):
    """Print a nicely formatted console summary."""
    total = len(df)
    resolved = total - resolution_counts.get("UNKNOWN", 0)
    pct = (resolved / total * 100) if total > 0 else 0

    print("\n" + "=" * 60)
    print("  Arena Leaderboard Enrichment Complete")
    print("=" * 60)
    print(f"Total models: {total}")
    print(f"Parameters resolved: {resolved} ({pct:.1f}%)")
    print(f"  - From AA cache:            {resolution_counts.get('artificial_analysis', 0)}")
    print(f"  - From overrides:           {resolution_counts.get('override', 0)}")
    print(f"  - From name parsing:        {resolution_counts.get('name_parsing', 0)}")
    print(f"  - From AA single scrape:    {resolution_counts.get('aa_single_scrape', 0)}")
    print(f"  - Unresolved:               {resolution_counts.get('UNKNOWN', 0)}")

    for precision in ["BF16", "FP8"]:
        p_lower = precision.lower()
        serving_col = f"vram_{p_lower}_serving_gb"
        if serving_col not in df.columns:
            continue

        print(f"\n=== Best Model Per GPU ({precision} serving) ===")
        for gpu_key, gpu_vram in GPUS_BY_VRAM:
            display = GPU_DISPLAY_NAMES[gpu_key]
            fits = df[df[serving_col].notna() & (df[serving_col] <= gpu_vram)]
            if not fits.empty:
                best = fits.iloc[0]
                rank = int(best["rank"]) if pd.notna(best["rank"]) else "?"
                name = best["model_name"]
                total_b = best.get("total_params_b", "?")
                arch = best.get("architecture", "?")
                serving = round(best[serving_col], 1)
                arch_label = f" {arch.upper()}" if arch and arch != "dense" else ""
                print(
                    f"  {display:20s} ({gpu_vram:3d} GB): "
                    f"#{rank:<4} {name} ({total_b}B{arch_label}, ~{serving} GB serving)"
                )
            else:
                print(f"  {display:20s} ({gpu_vram:3d} GB): No model fits")

    print()


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Enrich Arena.ai leaderboard with VRAM estimates"
    )
    parser.add_argument(
        "--input", "-i",
        help="Path to a pre-downloaded arena leaderboard CSV",
        default=None,
    )
    parser.add_argument(
        "--no-network",
        action="store_true",
        help="Skip network-based resolution (AA). Use only overrides + name parsing.",
    )
    parser.add_argument(
        "--update-readme",
        action="store_true",
        help="Update the root README.md with leaderboard tables.",
    )
    parser.add_argument(
        "--refresh-cache",
        action="store_true",
        help="Force re-fetch of the Artificial Analysis model cache.",
    )
    parser.add_argument(
        "--output-dir", "-o",
        help="Output directory (default: arena_enrichment/output/)",
        default=OUTPUT_DIR,
    )
    args = parser.parse_args()

    # Step 1: Get the leaderboard data
    print("Step 1: Loading Arena.ai leaderboard...")
    try:
        df = scrape_arena_leaderboard(input_csv=args.input)
    except RuntimeError as e:
        print(f"\nError: {e}")
        sys.exit(1)

    print(f"  Loaded {len(df)} models")

    # Ensure sorted by rank
    if "rank" in df.columns and df["rank"].notna().any():
        df = df.sort_values("rank").reset_index(drop=True)

    # Step 2: Resolve parameter counts
    print("\nStep 2: Resolving model parameters...")
    use_network = not args.no_network

    # Pre-load AA model cache (one network request, or from disk)
    aa_lookup = None
    aa_is_stale = False
    if use_network:
        print("  Loading Artificial Analysis model database...")
        aa_lookup, aa_is_stale = get_aa_models(force_refresh=args.refresh_cache)
        print(f"  AA database: {len(aa_lookup)} entries loaded")
        if aa_is_stale:
            print("  WARNING: Using stale AA cache (RSC fetch failed)")

    resolution_log = []
    resolution_counts = {}

    for idx, row in df.iterrows():
        model_name = row["model_name"]
        if pd.isna(model_name):
            continue

        params, source = resolve_model_params(
            model_name, use_network=use_network, aa_lookup=aa_lookup
        )
        resolution_counts[source] = resolution_counts.get(source, 0) + 1

        if params:
            df.at[idx, "total_params_b"] = params["total_params_b"]
            df.at[idx, "active_params_b"] = params["active_params_b"]
            df.at[idx, "architecture"] = params["architecture"]
        else:
            df.at[idx, "total_params_b"] = None
            df.at[idx, "active_params_b"] = None
            df.at[idx, "architecture"] = None
        df.at[idx, "resolution_source"] = source

        resolution_log.append({
            "model_name": model_name,
            "total_params_b": params["total_params_b"] if params else None,
            "active_params_b": params["active_params_b"] if params else None,
            "architecture": params["architecture"] if params else None,
            "resolution_source": source,
        })

        if (idx + 1) % 20 == 0 or idx == len(df) - 1:
            print(f"  Resolved {idx + 1}/{len(df)} models...", end="\r")

    print(f"  Resolved {len(df)}/{len(df)} models      ")

    # Step 3: Compute VRAM and GPU fit columns
    print("\nStep 3: Computing VRAM estimates and GPU fit...")
    df = add_all_vram_and_gpu_columns(df)
    print("  Done")

    # Step 4: Output
    print("\nStep 4: Writing output files...")
    os.makedirs(args.output_dir, exist_ok=True)

    csv_path = os.path.join(args.output_dir, "arena_leaderboard_enriched.csv")
    df.to_csv(csv_path, index=False)
    print(f"  CSV: {csv_path}")

    xlsx_path = os.path.join(args.output_dir, "arena_leaderboard_enriched.xlsx")
    write_xlsx(df, resolution_log, xlsx_path)
    print(f"  XLSX: {xlsx_path}")

    # Step 5: Update README if requested
    if args.update_readme:
        print("\nStep 5: Updating README.md...")
        readme_content = generate_readme(df, resolution_counts, aa_is_stale=aa_is_stale)
        readme_path = os.path.join(REPO_ROOT, "README.md")
        with open(readme_path, "w") as f:
            f.write(readme_content)
        print(f"  README: {readme_path}")

    # Console summary
    print_console_summary(df, resolution_counts)

    # List unresolved models
    unresolved = [r for r in resolution_log if r["resolution_source"] == "UNKNOWN"]
    if unresolved:
        print(f"Unresolved models ({len(unresolved)}):")
        for r in unresolved:
            print(f"  - {r['model_name']}")


if __name__ == "__main__":
    main()
